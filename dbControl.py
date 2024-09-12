from openpyxl import Workbook, load_workbook
import os

class DB:
    def __init__(self):
        self.DB = "DBusers.xlsx"
        if not os.path.isfile(self.DB):
            wb = Workbook()
            wb.save(self.DB)

        
    def createUser(self, userName, password, mail):
        data = [userName, mail, password]
        wb = load_workbook(self.DB)
        ws = wb.active
        column = 1
        row = 1
        while True:
            if ws.cell(row=row, column=column).value == userName: return 0
            elif ws.cell(row=row, column=column).value == None: break
            row += 1
        while True:
            if ws.cell(row=row, column=column).value == None:
                for index, i in enumerate(data):
                    ws.cell(row=row, column=column+index).value = i
                break
            row += 1
        wb.save(self.DB)
        return 1

    
    def checkValidateUser(self, userName, password):
        wb = load_workbook(self.DB)
        ws = wb.active
        column = 1
        row = 1
        while True:
            if ws.cell(row=row, column=column).value == userName:
                if ws.cell(row=row, column=column+2).value == password: return 1
                else: return 0
            else: return 0